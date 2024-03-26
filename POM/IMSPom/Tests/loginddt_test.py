import pytest
import time
import openpyxl 
from Utils.driver import get_driver
from Utils.Urlconfig import URLConfig
from Pages.loginpage import LoginPage

@pytest.fixture(params=["chrome"])
def browser(request):
    return request.param

@pytest.fixture
def setup(browser):
    global driver
    driver = get_driver(browser)
    yield driver
    driver.close()

def readData():
    data_list = []
    path = "D:\QA Intern\POM\IMSPom\TestData\login.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']

    rows = sheet.max_row

    for r in range(2, rows+1):
        username = sheet.cell(row=r, column=1).value
        password = sheet.cell(row=r, column=2).value
        data_list.append((username, password))

    return data_list


@pytest.mark.parametrize("username, password", readData())
def test_login(setup, username, password):
    driver = setup
    login_page = LoginPage(driver)
    login_page.open_page(URLConfig.LOGIN_PAGE_URL)
    login_page.enter_username(username)
    login_page.enter_password(password)
    time.sleep(2)
    login_page.click_login()
    time.sleep(5)
    assert driver.title == "Projects | Infrastructure Management System Stage"

