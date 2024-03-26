import openpyxl
import csv

def read_data():
    data_list = []
    path = "D:\QA Intern\POM\IMSPom\TestData\login.xlsx"

    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']

    rows = sheet.max_row

    for r in range(2, rows+1):
        username = sheet.cell(row=r, column=1).value
        password = sheet.cell(row=r, column=2).value
        data_list.append((username, password))

    return data_list

def readData():
    data_csv = []
    csv_file = "D:\QA Intern\POM\IMSPom\TestData\login.csv"

    with open(csv_file, newline='') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        for row in reader:
            data_csv.append((row[0], row[1]))  

    return data_csv